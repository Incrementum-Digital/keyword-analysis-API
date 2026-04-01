"""
Bulk Sheet Exporter

Ported from CampaignForge bulkSheetExporter.ts
Generates Amazon SP bulk sheet XLSX files for campaign upload.
"""
from dataclasses import dataclass, field
from typing import Dict, List, Optional, Set
from decimal import Decimal
from io import BytesIO
import openpyxl
from openpyxl import Workbook


# All auto targeting types
ALL_AUTO_TARGETS = ['close_match', 'loose_match', 'substitutes', 'complements']

AUTO_TARGET_EXPRESSIONS: Dict[str, str] = {
    'close_match': 'close-match',
    'loose_match': 'loose-match',
    'substitutes': 'substitutes',
    'complements': 'complements',
}

# Amazon SP Bulk Sheet column order
COLUMNS = [
    'Product',
    'Entity',
    'Operation',
    'Campaign Id',
    'Ad Group Id',
    'Portfolio Id',
    'Ad Id (Read only)',
    'Keyword Id (Read only)',
    'Product Targeting Id (Read only)',
    'Campaign Name',
    'Ad Group Name',
    'Start Date',
    'End Date',
    'Targeting Type',
    'State',
    'Daily Budget',
    'SKU',
    'ASIN',
    'Ad Group Default Bid',
    'Bid',
    'Keyword Text',
    'Match Type',
    'Bidding Strategy',
    'Placement',
    'Percentage',
    'Product Targeting Expression',
]


@dataclass
class Keyword:
    """Keyword data for export."""
    id: str
    normalized_text: str
    original_text: str = ""
    search_volume: int = 0


@dataclass
class Campaign:
    """Campaign data for export."""
    id: str
    name: str
    match_type: str
    keyword_ids: List[str]
    daily_budget: Decimal
    default_bid: Decimal
    keyword_bid: Decimal
    bidding_strategy: str
    start_date: str
    status: str
    is_solo: bool = False
    is_auto: bool = False
    root_group: Optional[str] = None
    placement_multipliers_enabled: bool = False
    placement_top_of_search: int = 0
    placement_rest_of_search: int = 0
    placement_product_page: int = 0


@dataclass
class CampaignOverride:
    """Per-campaign overrides."""
    daily_budget: Optional[Decimal] = None
    paused_keyword_ids: Set[str] = field(default_factory=set)
    keyword_bids: Dict[str, Decimal] = field(default_factory=dict)


@dataclass
class CampaignNegatives:
    """Negative keywords for a campaign."""
    exact: List[str] = field(default_factory=list)
    phrase: List[str] = field(default_factory=list)


@dataclass
class ExportOptions:
    """Options for bulk sheet export."""
    include_campaign_rows: bool = True
    include_ad_group_rows: bool = True
    include_keyword_rows: bool = True
    include_product_ad_rows: bool = True
    sku: str = ""
    account_type: str = "seller"  # "seller" or "vendor"
    format: str = "new"  # "new" or "legacy"


def format_bidding_strategy(strategy: str) -> str:
    """Format bidding strategy for Amazon bulk sheet."""
    strategies = {
        'Fixed': 'Fixed bid',
        'Dynamic Down': 'Dynamic bids - down only',
        'Dynamic Up & Down': 'Dynamic bids - up and down',
    }
    return strategies.get(strategy, strategy)


def format_date(date_str: str) -> str:
    """Convert date from YYYY-MM-DD to YYYYMMDD format."""
    return date_str.replace('-', '') if date_str else ''


def calculate_base_bid(max_bid: Decimal, bidding_strategy: str, placement_enabled: bool,
                       top_pct: int = 0, rest_pct: int = 0, product_pct: int = 0) -> Decimal:
    """
    Calculate the base bid written to the bulk sheet from the user's Max Bid.

    Accounts for two amplification layers:
    1. Bidding Strategy: "Up and Down" can increase bids up to 100% (divisor=2)
    2. Placement Multipliers: further amplify by highest percentage

    Formula: base_bid = max_bid / strategy_divisor / (1 + highest_placement_pct / 100)
    """
    strategy_divisor = Decimal('2') if bidding_strategy == 'Dynamic Up & Down' else Decimal('1')
    strategy_adjusted = max_bid / strategy_divisor

    if placement_enabled:
        highest = max(top_pct, rest_pct, product_pct)
        if highest > 0:
            base = strategy_adjusted / (1 + Decimal(highest) / 100)
            return base.quantize(Decimal('0.01'))

    return strategy_adjusted.quantize(Decimal('0.01'))


def empty_row() -> Dict[str, str]:
    """Create an empty row with all columns."""
    return {col: '' for col in COLUMNS}


def generate_bulk_sheet(
    campaigns: List[Campaign],
    keywords: List[Keyword],
    overrides: Dict[str, CampaignOverride],
    options: ExportOptions,
    auto_targeting_groups: Optional[Dict[str, List[str]]] = None,
    campaign_negatives: Optional[Dict[str, CampaignNegatives]] = None,
) -> Workbook:
    """
    Generate an Amazon SP bulk sheet workbook.

    Args:
        campaigns: List of campaigns to export
        keywords: List of all keywords (for lookup)
        overrides: Per-campaign overrides
        options: Export options
        auto_targeting_groups: Auto targeting type selections by campaign root
        campaign_negatives: Negative keywords per campaign

    Returns:
        openpyxl Workbook ready for saving
    """
    rows: List[Dict[str, str]] = []
    keyword_map = {kw.id: kw for kw in keywords}
    product_ids = [s.strip() for s in options.sku.split(',') if s.strip()]
    is_vendor = options.account_type == 'vendor'

    for campaign in campaigns:
        override = overrides.get(campaign.id, CampaignOverride())
        daily_budget = override.daily_budget if override.daily_budget else campaign.daily_budget

        # Calculate base bid from max bid using placement multipliers and bidding strategy
        base_bid = calculate_base_bid(
            max_bid=campaign.keyword_bid,
            bidding_strategy=campaign.bidding_strategy,
            placement_enabled=campaign.placement_multipliers_enabled,
            top_pct=campaign.placement_top_of_search,
            rest_pct=campaign.placement_rest_of_search,
            product_pct=campaign.placement_product_page,
        )

        # 1. Campaign row
        if options.include_campaign_rows:
            row = empty_row()
            row['Product'] = 'Sponsored Products'
            row['Entity'] = 'Campaign'
            row['Operation'] = 'Create'
            row['Campaign Id'] = campaign.name
            row['Campaign Name'] = campaign.name
            row['Start Date'] = format_date(campaign.start_date)
            row['Targeting Type'] = 'Auto' if campaign.is_auto else 'Manual'
            row['State'] = campaign.status.lower()
            row['Daily Budget'] = str(daily_budget)
            row['Bidding Strategy'] = format_bidding_strategy(campaign.bidding_strategy)
            rows.append(row)

            # Placement multiplier rows (one per placement type with non-zero %)
            if campaign.placement_multipliers_enabled:
                for placement_type, pct in [
                    ('Placement Top', campaign.placement_top_of_search),
                    ('Placement Rest Of Search', campaign.placement_rest_of_search),
                    ('Placement Product Page', campaign.placement_product_page),
                ]:
                    if pct > 0:
                        prow = empty_row()
                        prow['Product'] = 'Sponsored Products'
                        prow['Entity'] = 'Bidding Adjustment'
                        prow['Operation'] = 'Create'
                        prow['Campaign Id'] = campaign.name
                        prow['Campaign Name'] = campaign.name
                        prow['Placement'] = placement_type
                        prow['Percentage'] = str(pct)
                        rows.append(prow)

        # 2. Ad Group row
        if options.include_ad_group_rows:
            row = empty_row()
            row['Product'] = 'Sponsored Products'
            row['Entity'] = 'Ad Group'
            row['Operation'] = 'Create'
            row['Campaign Id'] = campaign.name
            row['Ad Group Id'] = campaign.name
            row['Campaign Name'] = campaign.name
            row['Ad Group Name'] = campaign.name
            row['State'] = 'enabled'
            row['Ad Group Default Bid'] = str(campaign.default_bid)
            rows.append(row)

        # 3. Product Ad rows (one per SKU/ASIN)
        if options.include_product_ad_rows and product_ids:
            for pid in product_ids:
                row = empty_row()
                row['Product'] = 'Sponsored Products'
                row['Entity'] = 'Product Ad'
                row['Operation'] = 'Create'
                row['Campaign Id'] = campaign.name
                row['Ad Group Id'] = campaign.name
                row['Campaign Name'] = campaign.name
                row['Ad Group Name'] = campaign.name
                row['State'] = 'enabled'
                if is_vendor:
                    row['ASIN'] = pid
                else:
                    row['SKU'] = pid
                rows.append(row)

        # 4. Keyword rows (manual keyword campaigns: Exact, Phrase, Broad)
        if options.include_keyword_rows and not campaign.is_auto and campaign.match_type.lower() != 'product':
            paused_kw_ids = override.paused_keyword_ids
            for kw_id in campaign.keyword_ids:
                kw = keyword_map.get(kw_id)
                if not kw:
                    continue

                # Use per-keyword override bid if set, otherwise use calculated base bid
                bid = override.keyword_bids.get(kw_id, base_bid)
                is_paused = kw_id in paused_kw_ids

                row = empty_row()
                row['Product'] = 'Sponsored Products'
                row['Entity'] = 'Keyword'
                row['Operation'] = 'Create'
                row['Campaign Id'] = campaign.name
                row['Ad Group Id'] = campaign.name
                row['Campaign Name'] = campaign.name
                row['Ad Group Name'] = campaign.name
                row['State'] = 'paused' if is_paused else 'enabled'
                row['Bid'] = str(bid)
                row['Keyword Text'] = kw.normalized_text
                row['Match Type'] = campaign.match_type.lower()
                rows.append(row)

        # 5. Product Targeting rows (Product campaigns - ASIN targeting)
        if options.include_keyword_rows and campaign.match_type.lower() == 'product':
            for kw_id in campaign.keyword_ids:
                kw = keyword_map.get(kw_id)
                if not kw:
                    continue

                asin_text = (kw.original_text or kw.normalized_text).upper()

                row = empty_row()
                row['Product'] = 'Sponsored Products'
                row['Entity'] = 'Product Targeting'
                row['Operation'] = 'Create'
                row['Campaign Id'] = campaign.name
                row['Ad Group Id'] = campaign.name
                row['Campaign Name'] = campaign.name
                row['Ad Group Name'] = campaign.name
                row['State'] = 'enabled'
                row['Bid'] = str(campaign.keyword_bid)
                row['Product Targeting Expression'] = f'asin="{asin_text}"'
                rows.append(row)

        # 6. Auto campaign targeting type rows
        if campaign.is_auto:
            root_group = campaign.root_group or ''
            selected_types = set(
                (auto_targeting_groups or {}).get(root_group, ALL_AUTO_TARGETS)
            )
            for type_id in ALL_AUTO_TARGETS:
                row = empty_row()
                row['Product'] = 'Sponsored Products'
                row['Entity'] = 'Product Targeting'
                row['Operation'] = 'Create'
                row['Campaign Id'] = campaign.name
                row['Ad Group Id'] = campaign.name
                row['Campaign Name'] = campaign.name
                row['Ad Group Name'] = campaign.name
                row['State'] = 'enabled' if type_id in selected_types else 'paused'
                row['Bid'] = str(base_bid)
                row['Product Targeting Expression'] = AUTO_TARGET_EXPRESSIONS.get(type_id, type_id)
                rows.append(row)

        # 7. Campaign negative keyword rows
        # Look up negatives by campaign ID first, then by campaign name (for frontend-computed campaigns)
        negs = (campaign_negatives or {}).get(campaign.id) or (campaign_negatives or {}).get(campaign.name)
        if negs:
            for kw_text in negs.exact:
                row = empty_row()
                row['Product'] = 'Sponsored Products'
                row['Entity'] = 'Campaign Negative Keyword'
                row['Operation'] = 'Create'
                row['Campaign Id'] = campaign.name
                row['Campaign Name'] = campaign.name
                row['State'] = 'enabled'
                row['Keyword Text'] = kw_text
                row['Match Type'] = 'negativeExact'
                rows.append(row)

            for kw_text in negs.phrase:
                row = empty_row()
                row['Product'] = 'Sponsored Products'
                row['Entity'] = 'Campaign Negative Keyword'
                row['Operation'] = 'Create'
                row['Campaign Id'] = campaign.name
                row['Campaign Name'] = campaign.name
                row['State'] = 'enabled'
                row['Keyword Text'] = kw_text
                row['Match Type'] = 'negativePhrase'
                rows.append(row)

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = 'Sponsored Products Campaigns'

    # Write header
    ws.append(COLUMNS)

    # Column indices for date fields (1-based for openpyxl)
    start_date_col = COLUMNS.index('Start Date') + 1
    end_date_col = COLUMNS.index('End Date') + 1

    # Write data rows
    for row_data in rows:
        row_values = [row_data.get(col, '') for col in COLUMNS]
        ws.append(row_values)
        # Ensure Start Date and End Date cells are stored as text, not numbers.
        # openpyxl may coerce all-digit strings to float during append.
        row_idx = ws.max_row
        for col_idx in (start_date_col, end_date_col):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is not None and cell.value != '':
                cell.value = str(int(cell.value)) if isinstance(cell.value, (int, float)) else str(cell.value)
                cell.number_format = '@'  # Text format

    return wb


def workbook_to_bytes(wb: Workbook) -> bytes:
    """Convert workbook to bytes for HTTP response."""
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def get_export_summary(
    campaigns: List[Campaign],
    keywords: List[Keyword],
    campaign_negatives: Optional[Dict[str, CampaignNegatives]] = None,
) -> Dict:
    """
    Get summary statistics for export.

    Args:
        campaigns: List of campaigns
        keywords: List of keywords
        campaign_negatives: Negative keywords per campaign

    Returns:
        Dict with summary statistics
    """
    total_keywords = sum(len(c.keyword_ids) for c in campaigns if not c.is_auto)
    total_negatives = sum(
        len(n.exact) + len(n.phrase)
        for n in (campaign_negatives or {}).values()
    )

    # Match type breakdown
    match_type_breakdown: Dict[str, int] = {}
    for campaign in campaigns:
        mt = campaign.match_type.lower()
        match_type_breakdown[mt] = match_type_breakdown.get(mt, 0) + 1

    # Estimate total rows
    total_rows = (
        len(campaigns)  # Campaign rows
        + len(campaigns)  # Ad group rows
        + total_keywords  # Keyword rows
        + total_negatives  # Negative rows
        + sum(4 for c in campaigns if c.is_auto)  # Auto targeting rows
    )

    return {
        'total_campaigns': len(campaigns),
        'total_keywords': total_keywords,
        'total_negatives': total_negatives,
        'total_rows': total_rows,
        'match_type_breakdown': match_type_breakdown,
    }
